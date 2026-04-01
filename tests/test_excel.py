import io
from unittest.mock import patch
import openpyxl
from excel import procesar_excel


def _make_workbook(rows: list) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A13"] = "CUIT\nDonante/\nDonatario"
    ws["B13"] = "Denominación"
    for i, (cuit, denom) in enumerate(rows):
        row = 15 + i
        ws.cell(row=row, column=1, value=cuit)
        ws.cell(row=row, column=2, value=denom)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _mock_resolve(cuits, max_workers=12):
    return {c: "PEREZ JUAN" for c in cuits}


def test_fills_empty_denominacion():
    file_bytes = _make_workbook([(20123456789, None)])
    with patch("excel.resolve_cuits_parallel", side_effect=_mock_resolve):
        result = procesar_excel(file_bytes)
    wb = openpyxl.load_workbook(io.BytesIO(result))
    ws = wb.active
    assert ws.cell(15, 2).value == "PEREZ JUAN"


def test_does_not_overwrite_existing_denominacion():
    file_bytes = _make_workbook([(30712505873, "CIGARS SONS S.R.L.")])
    with patch("excel.resolve_cuits_parallel") as mock_r:
        result = procesar_excel(file_bytes)
    mock_r.assert_not_called()
    wb = openpyxl.load_workbook(io.BytesIO(result))
    ws = wb.active
    assert ws.cell(15, 2).value == "CIGARS SONS S.R.L."


def test_skips_row_with_empty_cuit():
    file_bytes = _make_workbook([(None, None)])
    with patch("excel.resolve_cuits_parallel") as mock_r:
        result = procesar_excel(file_bytes)
    mock_r.assert_not_called()


def test_converts_integer_cuit_to_string():
    file_bytes = _make_workbook([(20123456789, None)])
    with patch("excel.resolve_cuits_parallel") as mock_r:
        mock_r.return_value = {"20123456789": "GARCIA"}
        procesar_excel(file_bytes)
    mock_r.assert_called_once()
    assert mock_r.call_args[0][0] == ["20123456789"]


def test_preserves_other_columns():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A13"] = "CUIT\nDonante/\nDonatario"
    ws["B13"] = "Denominación"
    ws.cell(15, 1, 20123456789)
    ws.cell(15, 2, None)
    ws.cell(15, 3, "Valor columna C")
    ws.cell(15, 4, 42)
    buf = io.BytesIO()
    wb.save(buf)
    file_bytes = buf.getvalue()

    with patch("excel.resolve_cuits_parallel", side_effect=_mock_resolve):
        result = procesar_excel(file_bytes)

    wb2 = openpyxl.load_workbook(io.BytesIO(result))
    ws2 = wb2.active
    assert ws2.cell(15, 3).value == "Valor columna C"
    assert ws2.cell(15, 4).value == 42
