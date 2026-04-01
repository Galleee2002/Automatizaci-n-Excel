import io
import base64
import logging
import time
import openpyxl
from scraper import PARALLEL_WORKERS, resolve_cuits_parallel

logger = logging.getLogger(__name__)
logging.basicConfig(level=logging.DEBUG, format="%(asctime)s [%(levelname)s] %(message)s")

DATA_START_ROW = 15
CUIT_COL = 1
DENOM_COL = 2


def _cuit_str(val) -> str:
    raw = str(val).strip()
    if raw.endswith(".0") and raw[:-2].isdigit():
        return raw[:-2]
    return raw


def procesar_excel(file_bytes: bytes) -> bytes:
    t0 = time.perf_counter()
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
    logger.info("openpyxl.load_workbook: %.3f s", time.perf_counter() - t0)
    ws = wb.active

    logger.debug("Hoja activa: %s | max_row reportado: %d", ws.title, ws.max_row)

    pendientes: list[tuple[int, str]] = []
    for row in range(DATA_START_ROW, ws.max_row + 1):
        cuit_val = ws.cell(row, CUIT_COL).value
        denom_val = ws.cell(row, DENOM_COL).value

        if not cuit_val:
            continue
        if denom_val and str(denom_val).strip():
            continue

        pendientes.append((row, _cuit_str(cuit_val)))

    if pendientes:
        unique = list(dict.fromkeys(c for _, c in pendientes))
        cache = resolve_cuits_parallel(unique)
        for row, cuit_str in pendientes:
            ws.cell(row, DENOM_COL).value = cache.get(cuit_str, "NO ENCONTRADO")

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def procesar_excel_progreso(file_bytes: bytes):
    t0 = time.perf_counter()
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
    carga_workbook_s = time.perf_counter() - t0
    logger.info("openpyxl.load_workbook: %.3f s", carga_workbook_s)
    yield {"tipo": "metricas", "carga_workbook_s": carga_workbook_s}
    ws = wb.active

    logger.debug("Hoja activa: '%s' | max_row: %d", ws.title, ws.max_row)

    filas: list[tuple[int, str]] = []
    for row in range(DATA_START_ROW, ws.max_row + 1):
        cuit_val = ws.cell(row, CUIT_COL).value
        denom_val = ws.cell(row, DENOM_COL).value
        if not cuit_val:
            continue
        if denom_val and str(denom_val).strip():
            logger.debug("Fila %d: CUIT %s ya tiene denominación '%s', se omite", row, cuit_val, denom_val)
            continue
        filas.append((row, _cuit_str(cuit_val)))

    total = len(filas)
    logger.debug("Filas a procesar: %d", total)

    if total:
        unique = list(dict.fromkeys(c for _, c in filas))
        t_net = time.perf_counter()
        cache = resolve_cuits_parallel(unique)
        logger.info(
            "Consultas CUIT: %d únicos en %.3f s (pool %d hilos)",
            len(unique),
            time.perf_counter() - t_net,
            min(PARALLEL_WORKERS, len(unique)),
        )
        for i, (row, cuit_str) in enumerate(filas):
            denom = cache.get(cuit_str, "NO ENCONTRADO")
            ws.cell(row, DENOM_COL).value = denom
            logger.debug("Fila %d: CUIT %s → '%s'", row, cuit_str, denom)
            yield {"tipo": "progreso", "actual": i + 1, "total": total, "cuit": cuit_str, "denom": denom}

    logger.debug("Total celdas escritas: %d", total)

    output = io.BytesIO()
    wb.save(output)
    file_size = output.tell()
    logger.debug("Archivo guardado: %d bytes", file_size)
    yield {"tipo": "listo", "archivo_b64": base64.b64encode(output.getvalue()).decode()}
